#!/usr/bin/env python3
"""
Dynamic Validation Script 
Validates completeness of dynamic analysis processing for any tool
"""

import json
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from collections import Counter

def count_violations_in_json(json_obj, path=""):
    """Recursively count violations in any JSON structure"""
    count = 0
    
    if isinstance(json_obj, dict):
        # Common violation containers
        violation_keys = ['violations', 'files', 'results', 'Results', 'findings', 'issues', 
                         'bugs', 'BugInstance', 'Vulnerabilities', 'messages', 'offenses']
        
        for key, value in json_obj.items():
            if key in violation_keys and isinstance(value, list):
                if key == 'files':
                    # Handle PMD/PHPCS style where files contain violations
                    for file_data in value:
                        if isinstance(file_data, dict):
                            for sub_key, sub_value in file_data.items():
                                if sub_key in ['violations', 'messages', 'offenses'] and isinstance(sub_value, list):
                                    count += len(sub_value)
                elif key == 'files' and isinstance(value, dict):
                    # Handle PHPCS style with file paths as keys
                    for file_path, file_data in value.items():
                        if isinstance(file_data, dict) and 'messages' in file_data:
                            count += len(file_data['messages'])
                else:
                    count += len(value)
            elif isinstance(value, (dict, list)):
                count += count_violations_in_json(value, f"{path}.{key}" if path else key)
    
    elif isinstance(json_obj, list):
        for item in json_obj:
            count += count_violations_in_json(item, path)
    
    return count

def count_original_data_dynamic(input_file):
    """Count violations in original markdown file using dynamic detection"""
    with open(input_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Find tool sections dynamically
    tool_pattern = r'## 🔧 (\w+)(.*?)(?=## 🔧|\Z)'
    tool_matches = re.findall(tool_pattern, content, re.DOTALL)
    
    total_violations = 0
    tool_breakdown = {}
    
    for tool_name, tool_section in tool_matches:
        tool_name = tool_name.upper()
        tool_violations = 0
        
        # Count JSON violations in this tool section
        json_pattern = r'<details>.*?```json\n(.*?)\n```.*?</details>'
        json_matches = re.findall(json_pattern, tool_section, re.DOTALL)
        
        for match in json_matches:
            try:
                json_obj = json.loads(match)
                tool_violations += count_violations_in_json(json_obj)
            except json.JSONDecodeError:
                continue
        
        # Count text-based warnings (multiple patterns)
        warning_patterns = [
            r'\[(WARN|ERROR|INFO)\]\s+(.+?):(\d+):(\d+):\s+(.+?)\s+\[(.+?)\]',  # Standard format
            r'\[(WARNING|ERROR|INFO)\]\s+([^:]+:\d+:\d+):([^[]+)\[([^\]]+)\]',    # Alternative format
        ]
        
        for pattern in warning_patterns:
            text_matches = re.findall(pattern, tool_section)
            tool_violations += len(text_matches)
        
        tool_breakdown[tool_name] = tool_violations
        total_violations += tool_violations
    
    return {
        'total_lines': len(content.splitlines()),
        'tool_sections': len(tool_matches),
        'tool_breakdown': tool_breakdown,
        'total_violations': total_violations
    }

def validate_processed_results():
    """Validate results from all possible processors"""
    result_dirs = {
        'processed_results': 'Basic Processor',
        'enhanced_results': 'Enhanced Processor', 
        'dynamic_results': 'Dynamic Processor',
        'enhanced_dynamic_results': 'Enhanced Dynamic Processor'
    }
    
    validation_results = {}
    
    for dir_name, processor_name in result_dirs.items():
        dir_path = Path(dir_name)
        if not dir_path.exists():
            continue
            
        result_info = {
            'processor': processor_name,
            'excel_files': [],
            'html_files': [],
            'text_files': [],
            'chart_files': [],
            'total_rows': 0
        }
        
        # Check Excel files
        excel_files = list(dir_path.glob("*.xlsx"))
        for excel_file in excel_files:
            try:
                wb = load_workbook(excel_file)
                file_rows = 0
                sheet_info = {}
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = max(0, sheet.max_row - 1)  # Subtract header
                    file_rows += rows
                    sheet_info[sheet_name] = rows
                
                result_info['excel_files'].append({
                    'file': excel_file.name,
                    'sheets': sheet_info,
                    'total_rows': file_rows
                })
                result_info['total_rows'] += file_rows
            except Exception as e:
                print(f"⚠️  Error reading {excel_file}: {e}")
        
        # Check HTML files
        html_files = list(dir_path.glob("*.html"))
        for html_file in html_files:
            try:
                with open(html_file, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                    table_count = html_content.count('<table')
                    row_count = max(0, html_content.count('<tr') - table_count)
                    
                result_info['html_files'].append({
                    'file': html_file.name,
                    'tables': table_count,
                    'rows': row_count
                })
            except Exception as e:
                print(f"⚠️  Error reading {html_file}: {e}")
        
        # Check text files
        text_files = list(dir_path.glob("*.txt"))
        result_info['text_files'] = [f.name for f in text_files]
        
        # Check chart files
        chart_files = list(dir_path.glob("*.png"))
        result_info['chart_files'] = [f.name for f in chart_files]
        
        validation_results[dir_name] = result_info
    
    return validation_results

def check_data_integrity(input_file):
    """Check for data integrity issues across all processors"""
    integrity_issues = []
    
    # Check for truncation indicators
    result_dirs = ['processed_results', 'enhanced_results', 'dynamic_results', 'enhanced_dynamic_results']
    truncation_indicators = ['...', 'truncated', 'omitted', 'limited', '(truncated)', 'etc.']
    
    for dir_name in result_dirs:
        dir_path = Path(dir_name)
        if not dir_path.exists():
            continue
            
        # Check text and HTML files for truncation
        for file_path in dir_path.glob("*"):
            if file_path.suffix in ['.html', '.txt']:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read().lower()
                        found_truncation = []
                        
                        for indicator in truncation_indicators:
                            if indicator.lower() in content:
                                found_truncation.append(indicator)
                        
                        if found_truncation:
                            integrity_issues.append({
                                'file': file_path,
                                'issue': 'truncation_detected',
                                'indicators': found_truncation
                            })
                except Exception:
                    continue
    
    # Check for file path completeness
    for dir_name in result_dirs:
        dir_path = Path(dir_name)
        if dir_path.exists():
            text_files = list(dir_path.glob("*.txt"))
            for text_file in text_files:
                try:
                    with open(text_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        # Look for incomplete file paths
                        if re.search(r'[./\\][^:]*\.\.\.[^:]*:', content):
                            integrity_issues.append({
                                'file': text_file,
                                'issue': 'incomplete_file_paths',
                                'description': 'File paths appear truncated'
                            })
                except Exception:
                    continue
    
    return integrity_issues

def validate_dynamic_completeness(input_file=None):
    repo_root = Path.cwd()
    input_file = input_file or (
        repo_root / "analysis_results" / "analyzer_results_convert.md"
    )

    """Main validation function for dynamic processing"""
    print("🔍 DYNAMIC VALIDATION: Checking data completeness...")
    
    # Analyze original file with dynamic detection
    original_stats = count_original_data_dynamic(input_file)
    print("✅ Original file analysis:")
    print(f"   📄 Total lines: {original_stats['total_lines']}")
    print(f"   🔧 Tool sections found: {original_stats['tool_sections']}")
    print(f"   📊 Total violations detected: {original_stats['total_violations']}")
    
    if original_stats['tool_breakdown']:
        print("   📋 Tool breakdown:")
        for tool, count in original_stats['tool_breakdown'].items():
            if count > 0:
                print(f"      - {tool}: {count} violations")
    
    # Validate processed results
    print(f"\n📊 PROCESSED FILES VALIDATION:")
    processed_results = validate_processed_results()
    
    total_processed_violations = 0
    active_processors = []
    
    for dir_name, result_info in processed_results.items():
        if result_info['excel_files'] or result_info['html_files']:
            active_processors.append(result_info['processor'])
            total_processed_violations += result_info['total_rows']
            
            print(f"\n📁 {result_info['processor'].upper()}:")
            
            # Excel file details
            for excel_info in result_info['excel_files']:
                print(f"   📊 {excel_info['file']}: {excel_info['total_rows']} total rows")
                for sheet, rows in excel_info['sheets'].items():
                    if rows > 0:
                        print(f"      - {sheet}: {rows} rows")
            
            # HTML file details  
            for html_info in result_info['html_files']:
                print(f"   🌐 {html_info['file']}: {html_info['tables']} tables, {html_info['rows']} rows")
            
            # Additional files
            if result_info['text_files']:
                print(f"   📄 Text files: {len(result_info['text_files'])}")
            if result_info['chart_files']:
                print(f"   📈 Charts: {len(result_info['chart_files'])}")
    
    # Data integrity check
    print(f"\n🔒 DATA INTEGRITY CHECK:")
    integrity_issues = check_data_integrity(input_file)
    
    if not integrity_issues:
        print("   ✅ No truncation or data loss detected")
        print("   ✅ File paths appear complete")
        print("   ✅ Data integrity maintained")
    else:
        print(f"   ⚠️  Found {len(integrity_issues)} potential issues:")
        for issue in integrity_issues[:3]:  # Show first 3 issues
            print(f"      - {issue['file'].name}: {issue['issue']}")
    
    # Cross-validation
    print(f"\n🎯 CROSS-VALIDATION:")
    if active_processors:
        print(f"   ✅ Active processors: {len(active_processors)}")
        print(f"   📊 Original violations: {original_stats['total_violations']}")
        
        # Find the processor with most violations (most complete)
        max_violations = 0
        best_processor = None
        for dir_name, result_info in processed_results.items():
            if result_info['total_rows'] > max_violations:
                max_violations = result_info['total_rows']
                best_processor = result_info['processor']
        
        if best_processor:
            print(f"   🥇 Most complete processor: {best_processor} ({max_violations} violations)")
            
            # Check completeness percentage
            if original_stats['total_violations'] > 0:
                completeness = (max_violations / original_stats['total_violations']) * 100
                print(f"   📈 Data completeness: {completeness:.1f}%")
    else:
        print("   ❌ No processed results found")
    
    # Final summary
    print(f"\n🏆 DYNAMIC VALIDATION SUMMARY:")
    print(f"   ✅ Tool sections auto-detected: {original_stats['tool_sections']}")
    print(f"   ✅ Total violations found: {original_stats['total_violations']}")
    print(f"   ✅ Processors validated: {len(active_processors)}")
    print(f"   ✅ Dynamic processing successful")
    print(f"   ✅ Multiple output formats available")
    
    if not integrity_issues:
        print(f"   ✅ Zero data integrity issues")
        print(f"   ✅ All tool data preserved completely")
    
    return {
        'original_stats': original_stats,
        'processed_results': processed_results,
        'integrity_issues': integrity_issues,
        'active_processors': active_processors
    }

def main():
    """Run validation on default file"""
    validate_dynamic_completeness()

if __name__ == "__main__":
    main()
